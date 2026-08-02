"""Load tabular data from txt/csv, xlsx/xls, json files into an in-memory SQLite database.

Each file becomes one table (first non-empty sheet for Excel files). Column names are
kept verbatim (Chinese/space allowed) and always double-quoted in SQL by consumers.
Table names are sanitized to [A-Za-z0-9_] so the SQL guard regex can match them.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

_SAFE_RE = re.compile(r"[^A-Za-z0-9_]")


def sanitize_table_name(name: str) -> str:
    cleaned = _SAFE_RE.sub("_", name)
    if not cleaned or cleaned[0].isdigit():
        cleaned = "t_" + cleaned
    return cleaned


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "gbk"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1")


def load_txt(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Parse a delimited text file; delimiter is auto-detected (comma/tab/semicolon/pipe)."""
    text = _decode_text(path.read_bytes())
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(text.splitlines(), dialect)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError(f"{path}: 文件为空或没有数据行")
    headers = [h.strip() if h is not None else "" for h in rows[0]]
    data = [[None if cell.strip() == "" else cell for cell in row] for row in rows[1:]]
    if not data:
        raise ValueError(f"{path}: 只有表头，没有数据行")
    return headers, _pad_rows(headers, data)


def load_xlsx(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Read the first non-empty sheet of a .xlsx workbook (values, not formulas)."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            rows = [list(row) for row in rows if any(cell not in (None, "") for cell in row)]
            if not rows:
                continue
            headers = [str(h) if h is not None else "" for h in rows[0]]
            data = rows[1:]
            if not data:
                raise ValueError(f"{path}: 工作表 '{sheet.title}' 只有表头，没有数据行")
            return headers, _pad_rows(headers, data)
        raise ValueError(f"{path}: 没有包含数据的工作表")
    finally:
        workbook.close()


def load_xls(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Read the first non-empty sheet of a legacy .xls workbook."""
    import xlrd

    workbook = xlrd.open_workbook(str(path))
    for sheet in workbook.sheets():
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        rows = [row for row in rows if any(str(cell).strip() for cell in row)]
        if not rows:
            continue
        headers = [str(h) if h not in (None, "") else "" for h in rows[0]]
        data = [[None if str(cell).strip() == "" else cell for cell in row] for row in rows[1:]]
        if not data:
            raise ValueError(f"{path}: 工作表 '{sheet.name}' 只有表头，没有数据行")
        return headers, _pad_rows(headers, data)
    raise ValueError(f"{path}: 没有包含数据的工作表")


def load_json(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Parse .json: list of objects -> table; object with an array of objects -> that array;
    plain object -> single-row table. Nested dicts are stringified."""
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    records: list[dict[str, Any]] | dict[str, Any] | None = None
    if isinstance(payload, list):
        if payload and all(isinstance(item, dict) for item in payload):
            records = payload
        else:
            raise ValueError(f"{path}: JSON 根数组元素必须是对象")
    elif isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
                records = value
                break
        else:
            records = payload
    else:
        raise ValueError(f"{path}: JSON 根必须是对象或对象数组")
    if records is None or (isinstance(records, list) and not records):
        raise ValueError(f"{path}: JSON 中没有可用的数据记录")

    if isinstance(records, dict):
        records = [records]

    def stringify(value: Any) -> Any:
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    headers: list[str] = []
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(key)
    data = [[stringify(record.get(header)) for header in headers] for record in records]
    return headers, data


def _infer_type(values: list[Any]) -> str:
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "TEXT"

    def all_int(items: list[Any]) -> bool:
        for item in items:
            if isinstance(item, bool):
                return False
            try:
                int(item)
            except (TypeError, ValueError):
                return False
        return True

    def all_float(items: list[Any]) -> bool:
        for item in items:
            if isinstance(item, bool):
                return False
            try:
                float(item)
            except (TypeError, ValueError):
                return False
        return True

    if all_int(non_null):
        return "INTEGER"
    if all_float(non_null):
        return "REAL"
    return "TEXT"


def _pad_rows(headers: list[str], rows: list[list[Any]]) -> list[list[Any]]:
    """Normalize ragged rows: pad short rows with None, truncate long rows."""
    width = len(headers)
    return [row[:width] + [None] * (width - len(row)) for row in rows]


def _unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for header in headers:
        name = header or "column"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        result.append(name)
    return result


_LOADERS = {".txt": load_txt, ".csv": load_txt, ".xlsx": load_xlsx, ".xls": load_xls, ".json": load_json}


def load_into_memory(files: list[Path]) -> sqlite3.Connection:
    """Parse every file and load the resulting tables into an in-memory SQLite database."""
    connection = sqlite3.connect(":memory:")
    used_names: dict[str, int] = {}
    for path in files:
        path = Path(path)
        loader = _LOADERS.get(path.suffix.lower())
        if loader is None:
            connection.close()
            raise ValueError(f"不支持的文件格式: {path.suffix or '(无扩展名)'}（支持: .db/.txt/.csv/.xlsx/.xls/.json）")
        headers, rows = loader(path)
        headers = _unique_headers(headers)
        types = [_infer_type([row[index] for row in rows]) for index in range(len(headers))]
        table_name = sanitize_table_name(path.stem)
        used_names[table_name] = used_names.get(table_name, 0) + 1
        if used_names[table_name] > 1:
            table_name = f"{table_name}_{used_names[table_name]}"

        columns = ", ".join(f'"{header}" {data_type}' for header, data_type in zip(headers, types))
        connection.execute(f'CREATE TABLE "{table_name}" ({columns})')
        placeholders = ", ".join("?" for _ in headers)
        connection.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', rows)
        connection.commit()
    return connection
